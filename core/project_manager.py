from pathlib import Path

from openpyxl import load_workbook


class ExcelManager:
    """
    Работа с книгой Excel.
    """

    TARGET_SHEET = "График по работам"

    def __init__(self):
        self.filename = ""
        self.workbook = None
        self.worksheet = None

    def open(self, filename: str):
        path = Path(filename)

        if not path.exists():
            raise FileNotFoundError(
                f"Файл не найден:\n{filename}"
            )

        self.filename = str(path)

        self.workbook = load_workbook(
            filename=self.filename,
            data_only=False,
        )

        if self.TARGET_SHEET not in self.workbook.sheetnames:
            raise RuntimeError(
                f"Лист '{self.TARGET_SHEET}' отсутствует."
            )

        self.worksheet = self.workbook[self.TARGET_SHEET]

    def close(self):
        self.workbook = None
        self.worksheet = None

    @property
    def opened(self):
        return self.workbook is not None

    @property
    def rows(self):
        return self.worksheet.max_row

    @property
    def columns(self):
        return self.worksheet.max_column

    def cell(self, row: int, column: int):
        """
        Возвращает значение ячейки.
        """
        return self.worksheet.cell(
            row=row,
            column=column,
        ).value

    def find_first_row(self):
        """
        Ищет первую строку техники.
        Предполагается, что:
        D = модель
        F = гаражный номер
        """

        for row in range(1, self.rows + 1):

            model = self.cell(row, 4)
            garage = self.cell(row, 6)

            if model and garage:
                return row

        raise RuntimeError(
            "Не удалось определить первую строку техники."
        )
    def find_by_garage_number(self, garage_number: str):
        """
        Возвращает список строк с указанным гаражным номером.
        """

        result = []

        start_row = self.find_first_row()

        for row in range(start_row, self.rows + 1):

            value = self.cell(row, 6)

            if value is None:
                continue

            if str(value).strip() == str(garage_number).strip():
                result.append(row)

        return result