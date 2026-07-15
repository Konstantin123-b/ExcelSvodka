from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta

from openpyxl.comments import Comment

from core.date_manager import DateManager
from core.equipment_manager import EquipmentManager
from core.models import MachineState, SvodkaRecord


class SvodkaLoader:
    """
    Загружает записи предыдущего дня из Excel.

    Читает все машины со статусами:

        >
        ав
        пл
        з

    и возвращает список SvodkaRecord.
    """

    def __init__(self, excel):

        self.excel = excel

        self.dates = DateManager(excel)

        self.equipment = EquipmentManager(excel)

    # ---------------------------------------------------------

    def load(self, date_string: str) -> list[SvodkaRecord]:

        previous_date = self._previous_date(date_string)

        column = self.dates.find(previous_date)

        records: list[SvodkaRecord] = []

        start_row = self.equipment.find_first_row()

        for row in range(start_row, self.excel.rows + 1):

            cell = self.excel.worksheet.cell(
                row=row,
                column=column,
            )

            code = str(cell.value or "").strip().lower()

            state = MachineState.from_code(code)

            if state is None:
                continue

            model = str(
                self.excel.cell(row, 4) or ""
            ).strip()

            garage = str(
                self.excel.cell(row, 6) or ""
            ).strip()

            description = ""
            employees = ""

            if cell.comment is not None:

                description, employees = self._parse_comment(
                    cell.comment
                )

            records.append(
                SvodkaRecord(
                    state=state,
                    garage_number=garage,
                    model=model,
                    description=description,
                    employees=employees,
                )
            )

        return records
      from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta

from openpyxl.comments import Comment

from core.date_manager import DateManager
from core.equipment_manager import EquipmentManager
from core.models import MachineState, SvodkaRecord


class SvodkaLoader:
    """
    Загружает записи предыдущего дня из Excel.

    Читает все машины со статусами:

        >
        ав
        пл
        з

    и возвращает список SvodkaRecord.
    """

    def __init__(self, excel):

        self.excel = excel

        self.dates = DateManager(excel)

        self.equipment = EquipmentManager(excel)

    # ---------------------------------------------------------

    def load(self, date_string: str) -> list[SvodkaRecord]:

        previous_date = self._previous_date(date_string)

        column = self.dates.find(previous_date)

        records: list[SvodkaRecord] = []

        start_row = self.equipment.find_first_row()

        for row in range(start_row, self.excel.rows + 1):

            cell = self.excel.worksheet.cell(
                row=row,
                column=column,
            )

            code = str(cell.value or "").strip().lower()

            state = MachineState.from_code(code)

            if state is None:
                continue

            model = str(
                self.excel.cell(row, 4) or ""
            ).strip()

            garage = str(
                self.excel.cell(row, 6) or ""
            ).strip()

            description = ""
            employees = ""

            if cell.comment is not None:

                description, employees = self._parse_comment(
                    cell.comment
                )

            records.append(
                SvodkaRecord(
                    state=state,
                    garage_number=garage,
                    model=model,
                    description=description,
                    employees=employees,
                )
            )

        return records
          # ---------------------------------------------------------

    @staticmethod
    def _sort_key(record: SvodkaRecord):

        return (
            record.model.lower(),
            record.garage_number.lower(),
        )

    # ---------------------------------------------------------

    @staticmethod
    def _remove_duplicates(
        records: list[SvodkaRecord],
    ) -> list[SvodkaRecord]:

        unique = {}

        for record in records:

            key = (
                record.garage_number.strip().lower(),
            )

            unique[key] = record

        return sorted(
            unique.values(),
            key=SvodkaLoader._sort_key,
        )

    # ---------------------------------------------------------

    def load_unique(
        self,
        date_string: str,
    ) -> list[SvodkaRecord]:
        """
        Загружает предыдущий день
        без повторов.
        """

        return self._remove_duplicates(
            self.load(date_string)
        )
