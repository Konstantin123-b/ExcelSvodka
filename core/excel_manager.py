from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.comments import Comment

from core.models import Equipment


class ExcelManager:
    """
    Центральный класс работы с Excel.
    """

    TARGET_SHEET = "График по работам"

    DATE_ROW = 9

    MODEL_COLUMN = 4

    GARAGE_COLUMN = 6

    def __init__(self):

        self.filename = ""

        self.workbook = None

        self.worksheet = None
            # ---------------------------------------------------------

    def open(
        self,
        filename: str,
    ):

        path = Path(filename)

        if not path.exists():

            raise FileNotFoundError(
                f"Файл не найден:\n{filename}"
            )

        self.filename = str(path)

        self.workbook = load_workbook(
            filename=self.filename,
            data_only=False,
        )

        if self.TARGET_SHEET not in self.workbook.sheetnames:

            raise RuntimeError(
                f"Лист '{self.TARGET_SHEET}' отсутствует."
            )

        self.worksheet = self.workbook[
            self.TARGET_SHEET
        ]

    # ---------------------------------------------------------

    def save(
        self,
        filename: str | None = None,
    ):

        if filename is None:
            filename = self.filename

        self.workbook.save(filename)

    # ---------------------------------------------------------

    def close(self):

        self.workbook = None

        self.worksheet = None

        self.filename = ""

    # ---------------------------------------------------------

    @property
    def opened(self):

        return self.workbook is not None

    # ---------------------------------------------------------

    @property
    def rows(self):

        return self.worksheet.max_row

    # ---------------------------------------------------------

    @property
    def columns(self):

        return self.worksheet.max_column
            # ---------------------------------------------------------

    def cell(
        self,
        row: int,
        column: int,
    ):

        return self.worksheet.cell(
            row=row,
            column=column,
        ).value

    # ---------------------------------------------------------

    def set_cell(
        self,
        row: int,
        column: int,
        value,
    ):

        self.worksheet.cell(
            row=row,
            column=column,
        ).value = value

    # ---------------------------------------------------------

    def set_fill(
        self,
        row: int,
        column: int,
        fill,
    ):

        self.worksheet.cell(
            row=row,
            column=column,
        ).fill = fill

    # ---------------------------------------------------------

    def set_comment(
        self,
        row: int,
        column: int,
        text: str,
        author: str = "ExcelSvodka",
    ):

        if not text:

            self.clear_comment(
                row,
                column,
            )

            return

        self.worksheet.cell(
            row=row,
            column=column,
        ).comment = Comment(
            text=text,
            author=author,
        )

    # ---------------------------------------------------------

    def clear_comment(
        self,
        row: int,
        column: int,
    ):

        self.worksheet.cell(
            row=row,
            column=column,
        ).comment = None
            # ---------------------------------------------------------

    def find_first_row(self) -> int:
        """
        Возвращает первую строку с техникой.
        """

        for row in range(1, self.rows + 1):

            model = self.cell(
                row,
                self.MODEL_COLUMN,
            )

            garage = self.cell(
                row,
                self.GARAGE_COLUMN,
            )

            if model and garage:
                return row

        raise RuntimeError(
            "Не удалось определить первую строку техники."
        )

    # ---------------------------------------------------------

    def iter_equipment(self):
        """
        Итератор по всей технике.
        """

        start = self.find_first_row()

        for row in range(
            start,
            self.rows + 1,
        ):

            model = self.cell(
                row,
                self.MODEL_COLUMN,
            )

            garage = self.cell(
                row,
                self.GARAGE_COLUMN,
            )

            if not garage:
                continue

            yield Equipment(
                row=row,
                model=str(model).strip(),
                garage_number=str(garage).strip(),
            )

    # ---------------------------------------------------------

    def find_by_garage_number(
        self,
        garage_number: str,
    ) -> list[Equipment]:

        garage_number = garage_number.strip()

        result = []

        for equipment in self.iter_equipment():

            if equipment.garage_number == garage_number:

                result.append(equipment)

        return result
            # ---------------------------------------------------------

    def find_date_column(
        self,
        date_string: str,
    ) -> int:
        """
        Возвращает номер столбца
        для указанной даты.
        """

        target = datetime.strptime(
            date_string,
            "%d.%m.%Y",
        ).date()

        for column in range(
            1,
            self.columns + 1,
        ):

            value = self.cell(
                self.DATE_ROW,
                column,
            )

            if value is None:
                continue

            if hasattr(value, "date"):

                if value.date() == target:
                    return column

        raise RuntimeError(
            f"Дата {date_string} не найдена."
        )

    # ---------------------------------------------------------

    def find_target_cell(
        self,
        garage_number: str,
        date_string: str,
    ) -> tuple[int, int]:
        """
        Возвращает координаты ячейки
        (row, column).
        """

        equipment = self.find_by_garage_number(
            garage_number
        )

        if not equipment:

            raise RuntimeError(
                f"Машина №{garage_number} не найдена."
            )

        if len(equipment) > 1:

            raise RuntimeError(
                f"Найдено несколько машин с гаражным номером {garage_number}."
            )

        column = self.find_date_column(
            date_string
        )

        return (
            equipment[0].row,
            column,
        )
            # ---------------------------------------------------------

    def equipment_exists(
        self,
        garage_number: str,
    ) -> bool:

        return len(
            self.find_by_garage_number(
                garage_number
            )
        ) > 0

    # ---------------------------------------------------------

    def get_equipment(
        self,
        garage_number: str,
    ) -> Equipment:

        equipment = self.find_by_garage_number(
            garage_number
        )

        if not equipment:

            raise RuntimeError(
                f"Машина №{garage_number} не найдена."
            )

        if len(equipment) > 1:

            raise RuntimeError(
                f"Найдено несколько машин с гаражным номером {garage_number}."
            )

        return equipment[0]

    # ---------------------------------------------------------

    def get_comment(
        self,
        row: int,
        column: int,
    ) -> str:

        comment = self.worksheet.cell(
            row=row,
            column=column,
        ).comment

        if comment is None:
            return ""

        return comment.text

    # ---------------------------------------------------------

    def clear_cell(
        self,
        row: int,
        column: int,
    ):

        cell = self.worksheet.cell(
            row=row,
            column=column,
        )

        cell.value = None

        cell.comment = None
