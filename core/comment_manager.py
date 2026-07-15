from openpyxl.comments import Comment

from core.models import SvodkaRecord


class CommentManager:
    """
    Работа с примечаниями Excel.
    """

    AUTHOR = "ExcelSvodka"

    def __init__(self, excel):

        self.excel = excel

    # ---------------------------------------------------------

    def get(self, row: int, column: int) -> str:

        cell = self.excel.worksheet.cell(
            row=row,
            column=column,
        )

        if cell.comment is None:
            return ""

        return cell.comment.text

    # ---------------------------------------------------------

    def clear(self, row: int, column: int):

        self.excel.worksheet.cell(
            row=row,
            column=column,
        ).comment = None

    # ---------------------------------------------------------

    def build(
        self,
        record: SvodkaRecord,
    ) -> str:
        """
        Формирует текст примечания.

        Формат:

        Замена двигателя

        Наработка: 18452 м/ч

        Иванов, Петров
        """

        parts = []

        description = record.description.strip()

        if description:
            parts.append(description)

        operating_hours = record.operating_hours.strip()

        if operating_hours:

            if parts:
                parts.append("")

            parts.append(
                f"Наработка: {operating_hours}"
            )

        employees = record.employees.strip()

        if employees:

            if parts:
                parts.append("")

            parts.append(employees)

        return "\n".join(parts)

    # ---------------------------------------------------------

    def set(
        self,
        row: int,
        column: int,
        record: SvodkaRecord,
    ):

        text = self.build(record)

        cell = self.excel.worksheet.cell(
            row=row,
            column=column,
        )

        if text:

            cell.comment = Comment(
                text=text,
                author=self.AUTHOR,
            )

        else:

            cell.comment = None

    # ---------------------------------------------------------

    def parse(
        self,
        row: int,
        column: int,
    ) -> tuple[str, str, str]:
        """
        Возвращает

        description,
        operating_hours,
        employees
        """

        text = self.get(
            row,
            column,
        ).replace(
            "\r\n",
            "\n",
        ).strip()

        if not text:
            return "", "", ""

        description = ""
        operating_hours = ""
        employees = ""

        blocks = [
            x.strip()
            for x in text.split("\n\n")
            if x.strip()
        ]

        for block in blocks:

            if block.startswith("Наработка:"):

                operating_hours = (
                    block.replace(
                        "Наработка:",
                        "",
                    )
                    .strip()
                )

                continue

            if "," in block:

                employees = block

                continue

            if not description:

                description = block

        return (
            description,
            operating_hours,
            employees,
        )
