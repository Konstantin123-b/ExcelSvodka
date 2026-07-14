from openpyxl.styles import PatternFill


class StyleManager:

    COLORS = {
        "ав": "FFFF0000",   # красный
        "пл": "FF8B4513",   # коричневый
        "з":  "FF800080",   # фиолетовый
    }

    @classmethod
    def get_fill(cls, code: str):

        code = code.strip().lower()

        if code not in cls.COLORS:
            return PatternFill(fill_type=None)

        color = cls.COLORS[code]

        return PatternFill(
            fill_type="solid",
            start_color=color,
            end_color=color,
        )